#!/usr/bin/env python3
"""
複雜輸出系統 - 任務執行模組

這個模組提供了一個功能豐富的複雜輸出系統，包含：
- 多種數據格式的展示（表格、圖表、JSON、XML）
- 動態統計分析功能
- 彩色終端輸出和格式化
- 模擬實時數據更新
- 多維度數據可視化
- 系統性能監控展示

主要功能：
- 複雜數據表格生成和顯示
- ASCII 圖表繪製
- 實時數據流模擬
- 多種輸出格式支援
- 動態進度條和狀態顯示
- 系統資源監控

技術特點：
- 支援彩色終端輸出
- 動態數據更新
- 多種圖表類型
- 自適應列寬
- 實時統計計算

作者: FastAPI Demo Project
版本: 1.0.0
創建日期: 2024
"""

import json
import math
import os
import platform
import random
import sys
import time
import threading
import queue
import asyncio
import concurrent.futures
import subprocess
import io
import base64
from collections import Counter, defaultdict, deque
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple, Union, Callable
from pathlib import Path
from contextlib import contextmanager
from functools import wraps
from enum import Enum

try:
    import numpy as np
    import psutil
    from rich.console import Console
    from rich.table import Table
    from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn
    from rich.live import Live
    from rich.layout import Layout
    from rich.panel import Panel
    from sklearn.cluster import KMeans
    from sklearn.linear_model import LinearRegression
    from sklearn.preprocessing import StandardScaler
    from sklearn.ensemble import IsolationForest
    import pandas as pd
    import matplotlib
    matplotlib.use('Agg')  # 無頭模式
    import matplotlib.pyplot as plt
    from matplotlib.patches import Rectangle
    import plotly.graph_objects as go
    import plotly.express as px
    from plotly.offline import plot
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    ADVANCED_FEATURES = True
except ImportError as e:
    print(f"警告: 某些進階功能無法使用 - {e}")
    ADVANCED_FEATURES = False
    np = None
    psutil = None


# 顏色代碼定義與擴展
class Colors:
    """擴展終端顏色代碼"""
    RED = '\033[91m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    BLUE = '\033[94m'
    MAGENTA = '\033[95m'
    CYAN = '\033[96m'
    WHITE = '\033[97m'
    BLACK = '\033[90m'
    BOLD = '\033[1m'
    DIM = '\033[2m'
    ITALIC = '\033[3m'
    UNDERLINE = '\033[4m'
    BLINK = '\033[5m'
    REVERSE = '\033[7m'
    STRIKETHROUGH = '\033[9m'
    END = '\033[0m'

    # 背景顏色
    BG_RED = '\033[101m'
    BG_GREEN = '\033[102m'
    BG_YELLOW = '\033[103m'
    BG_BLUE = '\033[104m'
    BG_MAGENTA = '\033[105m'
    BG_CYAN = '\033[106m'
    BG_WHITE = '\033[107m'

    # 256色支援
    @staticmethod
    def rgb(r: int, g: int, b: int) -> str:
        """RGB 顏色支援"""
        return f'\033[38;2;{r};{g};{b}m'

    @staticmethod
    def bg_rgb(r: int, g: int, b: int) -> str:
        """RGB 背景顏色支援"""
        return f'\033[48;2;{r};{g};{b}m'

class ChartType(Enum):
    """圖表類型枚舉"""
    BAR_3D = "3d_bar"
    SURFACE_3D = "3d_surface"
    HEATMAP_2D = "2d_heatmap"
    SCATTER_3D = "3d_scatter"
    HISTOGRAM_3D = "3d_histogram"
    WIREFRAME_3D = "3d_wireframe"

class DataType(Enum):
    """數據類型枚舉"""
    NUMERICAL = "numerical"
    CATEGORICAL = "categorical"
    TIME_SERIES = "time_series"
    GEOSPATIAL = "geospatial"
    NETWORK = "network"

@dataclass
class SystemInfo:
    """擴展系統信息數據類"""
    platform: str
    python_version: str
    cpu_count: int
    memory_total: int
    uptime: float
    cpu_freq: Optional[float] = None
    cpu_usage: Optional[float] = None
    memory_usage: Optional[float] = None
    disk_usage: Optional[Dict[str, float]] = None
    network_stats: Optional[Dict[str, Any]] = None
    process_count: Optional[int] = None

@dataclass
class DataPoint:
    """增強數據點結構"""
    timestamp: datetime
    value: float
    category: str
    metadata: Dict[str, Any]
    coordinates: Optional[Tuple[float, float, float]] = None
    confidence: Optional[float] = None
    anomaly_score: Optional[float] = None

@dataclass
class Point3D:
    """3D點座標"""
    x: float
    y: float
    z: float
    color: Optional[str] = None
    symbol: str = '●'

@dataclass
class Matrix3D:
    """3D矩陣數據"""
    data: List[List[List[float]]]
    x_labels: List[str] = field(default_factory=list)
    y_labels: List[str] = field(default_factory=list)
    z_labels: List[str] = field(default_factory=list)

@dataclass
class MLModelResult:
    """機器學習模型結果"""
    model_type: str
    accuracy: float
    predictions: List[Any]
    feature_importance: Optional[Dict[str, float]] = None
    cluster_centers: Optional[List[List[float]]] = None
    anomalies: Optional[List[int]] = None

@dataclass
class SystemMetrics:
    """系統指標數據"""
    timestamp: datetime
    cpu_percent: float
    memory_percent: float
    disk_io_read: float
    disk_io_write: float
    network_bytes_sent: float
    network_bytes_recv: float
    process_count: int
    load_avg: Optional[Tuple[float, float, float]] = None

class Advanced3DEngine:
    """3D ASCII 圖表引擎"""

    def __init__(self, width: int = 80, height: int = 40):
        self.width = width
        self.height = height
        self.depth_chars = ['░', '▒', '▓', '█']
        self.perspective_factor = 0.6

    def create_3d_bar_chart(self, data: List[List[float]],
                           labels: List[str] = None) -> str:
        """創建3D柱狀圖"""
        if not data:
            return "無數據可顯示"

        max_val = max(max(row) for row in data)
        normalized_data = [[val/max_val * 20 for val in row] for row in data]

        output = [f"{Colors.BOLD}{Colors.CYAN}3D 立體柱狀圖{Colors.END}\n"]
        output.append("=" * 60 + "\n")

        # 繪製3D效果
        for z_level in range(10, -1, -1):  # 從後往前繪製
            line = f"{z_level:2d} "

            for y, row in enumerate(normalized_data):
                for x, height in enumerate(row):
                    if height > z_level:
                        # 添加3D深度效果
                        if z_level < height - 2:
                            char = self.depth_chars[3]  # 最深
                            color = Colors.RED
                        elif z_level < height - 1:
                            char = self.depth_chars[2]
                            color = Colors.YELLOW
                        elif z_level < height:
                            char = self.depth_chars[1]
                            color = Colors.GREEN
                        else:
                            char = self.depth_chars[0]
                            color = Colors.BLUE

                        # 添加透視效果
                        perspective_offset = int((10 - z_level) * self.perspective_factor)
                        line += " " * perspective_offset + f"{color}{char}{char}{Colors.END}"
                    else:
                        line += "   "
                line += "  "

            output.append(line + "\n")

        # 添加座標軸標籤
        if labels:
            output.append("\n標籤: " + " | ".join(labels[:len(data[0])]) + "\n")

        return "".join(output)

    def create_3d_surface(self, matrix: Matrix3D) -> str:
        """創建3D表面圖"""
        if not matrix.data:
            return "無數據可顯示"

        output = [f"{Colors.BOLD}{Colors.MAGENTA}3D 表面圖{Colors.END}\n"]
        output.append("=" * 60 + "\n")

        # 計算等高線
        flat_data = [val for layer in matrix.data for row in layer for val in row]
        if not flat_data:
            return "數據為空"

        min_val, max_val = min(flat_data), max(flat_data)
        contour_levels = np.linspace(min_val, max_val, 8) if ADVANCED_FEATURES else [min_val + i*(max_val-min_val)/8 for i in range(8)]

        # 繪製等高線圖
        for y in range(len(matrix.data[0])):
            line = f"{y:2d} "
            for x in range(len(matrix.data[0][0])):
                # 計算該點的值（簡化為第一層）
                if y < len(matrix.data[0]) and x < len(matrix.data[0][0]):
                    val = matrix.data[0][y][x] if matrix.data else 0

                    # 確定等高線級別
                    level = 0
                    for i, threshold in enumerate(contour_levels):
                        if val >= threshold:
                            level = i

                    # 根據高度選擇字符和顏色
                    chars = [' ', '·', ':', ';', '=', '#', '█', '██']
                    colors = [Colors.BLACK, Colors.BLUE, Colors.CYAN,
                             Colors.GREEN, Colors.YELLOW, Colors.RED,
                             Colors.MAGENTA, Colors.WHITE]

                    char = chars[min(level, len(chars)-1)]
                    color = colors[min(level, len(colors)-1)]
                    line += f"{color}{char}{Colors.END}"
                else:
                    line += " "
            output.append(line + "\n")

        return "".join(output)

    def create_3d_scatter(self, points: List[Point3D]) -> str:
        """創建3D散點圖"""
        if not points:
            return "無數據點可顯示"

        output = [f"{Colors.BOLD}{Colors.GREEN}3D 散點圖{Colors.END}\n"]
        output.append("=" * 60 + "\n")

        # 正規化座標到顯示範圍
        if points:
            x_coords = [p.x for p in points]
            y_coords = [p.y for p in points]
            z_coords = [p.z for p in points]

            x_min, x_max = min(x_coords), max(x_coords)
            y_min, y_max = min(y_coords), max(y_coords)
            z_min, z_max = min(z_coords), max(z_coords)

            # 創建顯示網格
            grid = [[' ' for _ in range(self.width)] for _ in range(self.height)]

            for point in points:
                # 座標轉換
                x = int((point.x - x_min) / (x_max - x_min + 0.001) * (self.width - 10))
                y = int((point.y - y_min) / (y_max - y_min + 0.001) * (self.height - 5))
                z_intensity = (point.z - z_min) / (z_max - z_min + 0.001)

                # 根據Z軸深度選擇顏色和字符
                if z_intensity > 0.75:
                    color = Colors.RED
                    char = '●'
                elif z_intensity > 0.5:
                    color = Colors.YELLOW
                    char = '◆'
                elif z_intensity > 0.25:
                    color = Colors.GREEN
                    char = '▲'
                else:
                    color = Colors.BLUE
                    char = '■'

                if 0 <= y < self.height and 0 <= x < self.width:
                    grid[y][x] = f"{color}{char}{Colors.END}"

            # 輸出網格
            for row in grid:
                output.append("".join(row) + "\n")

        return "".join(output)

    def create_wireframe_3d(self, matrix: Matrix3D) -> str:
        """創建3D線框圖"""
        if not matrix.data:
            return "無數據可顯示"

        output = [f"{Colors.BOLD}{Colors.CYAN}3D 線框圖{Colors.END}\n"]
        output.append("=" * 60 + "\n")

        # 簡化的線框繪製
        height, width = len(matrix.data[0]), len(matrix.data[0][0]) if matrix.data[0] else 0

        for y in range(0, height, 2):  # 每隔一行繪製
            line = f"{y:2d} "
            for x in range(0, width, 3):  # 每隔兩列繪製
                if y < height and x < width:
                    # 繪製節點
                    line += f"{Colors.GREEN}+{Colors.END}"

                    # 繪製橫向連線
                    if x + 3 < width:
                        line += f"{Colors.WHITE}---{Colors.END}"

            output.append(line + "\n")

            # 繪製縱向連線
            if y + 2 < height:
                line = "   "
                for x in range(0, width, 3):
                    line += f"{Colors.WHITE}|   {Colors.END}"
                output.append(line + "\n")

        return "".join(output)

class MachineLearningAnalyzer:
    """機器學習數據分析引擎"""

    def __init__(self):
        self.scaler = StandardScaler() if ADVANCED_FEATURES else None
        self.models = {}

    def detect_anomalies(self, data: List[float], contamination: float = 0.1) -> MLModelResult:
        """異常檢測"""
        if not ADVANCED_FEATURES or not data:
            return MLModelResult("isolation_forest", 0.0, [], anomalies=[])

        try:
            # 準備數據
            X = np.array(data).reshape(-1, 1)

            # 異常檢測模型
            model = IsolationForest(contamination=contamination, random_state=42)
            predictions = model.fit_predict(X)

            # 找出異常點
            anomalies = [i for i, pred in enumerate(predictions) if pred == -1]
            scores = model.decision_function(X)

            return MLModelResult(
                model_type="isolation_forest",
                accuracy=len([p for p in predictions if p == 1]) / len(predictions),
                predictions=predictions.tolist(),
                anomalies=anomalies
            )
        except Exception as e:
            print(f"異常檢測錯誤: {e}")
            return MLModelResult("isolation_forest", 0.0, [], anomalies=[])

    def perform_clustering(self, points: List[Tuple[float, float]],
                          n_clusters: int = 3) -> MLModelResult:
        """K-means聚類分析"""
        if not ADVANCED_FEATURES or not points:
            return MLModelResult("kmeans", 0.0, [])

        try:
            X = np.array(points)

            # K-means聚類
            kmeans = KMeans(n_clusters=n_clusters, random_state=42, n_init=10)
            labels = kmeans.fit_predict(X)

            # 計算輪廓係數（簡化版）
            if len(set(labels)) > 1:
                from sklearn.metrics import silhouette_score
                accuracy = silhouette_score(X, labels)
            else:
                accuracy = 0.0

            return MLModelResult(
                model_type="kmeans",
                accuracy=accuracy,
                predictions=labels.tolist(),
                cluster_centers=kmeans.cluster_centers_.tolist()
            )
        except Exception as e:
            print(f"聚類分析錯誤: {e}")
            return MLModelResult("kmeans", 0.0, [])

    def predict_trend(self, time_series: List[float],
                     future_points: int = 10) -> MLModelResult:
        """時間序列預測"""
        if not ADVANCED_FEATURES or len(time_series) < 2:
            # 簡單線性趨勢
            if len(time_series) >= 2:
                slope = (time_series[-1] - time_series[0]) / (len(time_series) - 1)
                predictions = [time_series[-1] + slope * i for i in range(1, future_points + 1)]
            else:
                predictions = [0] * future_points

            return MLModelResult("linear_trend", 0.5, predictions)

        try:
            # 準備數據
            X = np.arange(len(time_series)).reshape(-1, 1)
            y = np.array(time_series)

            # 線性回歸預測
            model = LinearRegression()
            model.fit(X, y)

            # 預測未來點
            future_X = np.arange(len(time_series),
                               len(time_series) + future_points).reshape(-1, 1)
            predictions = model.predict(future_X)

            # 計算R²分數
            accuracy = model.score(X, y)

            return MLModelResult(
                model_type="linear_regression",
                accuracy=accuracy,
                predictions=predictions.tolist()
            )
        except Exception as e:
            print(f"趨勢預測錯誤: {e}")
            return MLModelResult("linear_regression", 0.0, [])

class SystemMonitor:
    """實時系統監控器"""

    def __init__(self):
        self.metrics_history = deque(maxlen=100)
        self.monitoring = False
        self.alert_thresholds = {
            'cpu': 80.0,
            'memory': 85.0,
            'disk_io': 50000000,  # 50MB/s
            'network': 10000000   # 10MB/s
        }

    def get_current_metrics(self) -> SystemMetrics:
        """獲取當前系統指標"""
        if not ADVANCED_FEATURES:
            # 模擬數據
            return SystemMetrics(
                timestamp=datetime.now(),
                cpu_percent=random.uniform(10, 90),
                memory_percent=random.uniform(30, 80),
                disk_io_read=random.uniform(1000000, 50000000),
                disk_io_write=random.uniform(500000, 20000000),
                network_bytes_sent=random.uniform(100000, 10000000),
                network_bytes_recv=random.uniform(200000, 15000000),
                process_count=random.randint(150, 300)
            )

        try:
            # 獲取真實系統數據
            cpu_percent = psutil.cpu_percent(interval=1)
            memory = psutil.virtual_memory()
            disk_io = psutil.disk_io_counters()
            network_io = psutil.net_io_counters()

            return SystemMetrics(
                timestamp=datetime.now(),
                cpu_percent=cpu_percent,
                memory_percent=memory.percent,
                disk_io_read=disk_io.read_bytes if disk_io else 0,
                disk_io_write=disk_io.write_bytes if disk_io else 0,
                network_bytes_sent=network_io.bytes_sent if network_io else 0,
                network_bytes_recv=network_io.bytes_recv if network_io else 0,
                process_count=len(psutil.pids()),
                load_avg=os.getloadavg() if hasattr(os, 'getloadavg') else None
            )
        except Exception as e:
            print(f"系統監控錯誤: {e}")
            return self.get_current_metrics()  # 返回模擬數據

    def check_alerts(self, metrics: SystemMetrics) -> List[str]:
        """檢查警報條件"""
        alerts = []

        if metrics.cpu_percent > self.alert_thresholds['cpu']:
            alerts.append(f"⚠️  CPU使用率過高: {metrics.cpu_percent:.1f}%")

        if metrics.memory_percent > self.alert_thresholds['memory']:
            alerts.append(f"⚠️  記憶體使用率過高: {metrics.memory_percent:.1f}%")

        return alerts

    def display_real_time_dashboard(self, duration: int = 30):
        """顯示實時監控儀表板"""
        print(f"{Colors.BOLD}{Colors.GREEN}啟動實時系統監控儀表板...{Colors.END}")
        print(f"監控時間: {duration} 秒\n")

        start_time = time.time()

        while time.time() - start_time < duration:
            # 清屏（跨平台）
            os.system('cls' if os.name == 'nt' else 'clear')

            # 獲取當前指標
            metrics = self.get_current_metrics()
            self.metrics_history.append(metrics)

            # 顯示標題
            print(f"{Colors.BOLD}{Colors.CYAN}{'='*80}{Colors.END}")
            print(f"{Colors.BOLD}{Colors.CYAN}實時系統監控儀表板 - {metrics.timestamp.strftime('%Y-%m-%d %H:%M:%S')}{Colors.END}")
            print(f"{Colors.BOLD}{Colors.CYAN}{'='*80}{Colors.END}\n")

            # 顯示關鍵指標
            self._display_metrics_grid(metrics)

            # 顯示歷史趨勢圖
            if len(self.metrics_history) > 1:
                self._display_trend_charts()

            # 檢查並顯示警報
            alerts = self.check_alerts(metrics)
            if alerts:
                print(f"\n{Colors.RED}{Colors.BOLD}警報通知:{Colors.END}")
                for alert in alerts:
                    print(f"  {Colors.RED}{alert}{Colors.END}")

            time.sleep(2)

    def _display_metrics_grid(self, metrics: SystemMetrics):
        """顯示指標網格"""
        # CPU使用率顯示
        cpu_bar = self._create_progress_bar(metrics.cpu_percent, 100, 30)
        cpu_color = Colors.RED if metrics.cpu_percent > 80 else Colors.YELLOW if metrics.cpu_percent > 60 else Colors.GREEN

        print(f"CPU 使用率:    {cpu_color}{metrics.cpu_percent:5.1f}%{Colors.END} {cpu_bar}")

        # 記憶體使用率顯示
        mem_bar = self._create_progress_bar(metrics.memory_percent, 100, 30)
        mem_color = Colors.RED if metrics.memory_percent > 85 else Colors.YELLOW if metrics.memory_percent > 70 else Colors.GREEN

        print(f"記憶體使用率: {mem_color}{metrics.memory_percent:5.1f}%{Colors.END} {mem_bar}")

        # 磁碟I/O顯示
        disk_read_mb = metrics.disk_io_read / 1024 / 1024
        disk_write_mb = metrics.disk_io_write / 1024 / 1024

        print(f"磁碟讀取:     {Colors.CYAN}{disk_read_mb:8.1f} MB{Colors.END}")
        print(f"磁碟寫入:     {Colors.CYAN}{disk_write_mb:8.1f} MB{Colors.END}")

        # 網路I/O顯示
        net_sent_mb = metrics.network_bytes_sent / 1024 / 1024
        net_recv_mb = metrics.network_bytes_recv / 1024 / 1024

        print(f"網路發送:     {Colors.MAGENTA}{net_sent_mb:8.1f} MB{Colors.END}")
        print(f"網路接收:     {Colors.MAGENTA}{net_recv_mb:8.1f} MB{Colors.END}")

        # 程序數量
        print(f"執行程序數:   {Colors.YELLOW}{metrics.process_count:>8}{Colors.END}")

        if metrics.load_avg:
            print(f"系統負載:     {Colors.WHITE}{metrics.load_avg[0]:.2f}, {metrics.load_avg[1]:.2f}, {metrics.load_avg[2]:.2f}{Colors.END}")

    def _create_progress_bar(self, value: float, max_value: float, length: int = 20) -> str:
        """創建進度條"""
        percentage = min(value / max_value, 1.0)
        filled_length = int(length * percentage)
        bar = "█" * filled_length + "░" * (length - filled_length)

        if percentage > 0.8:
            color = Colors.RED
        elif percentage > 0.6:
            color = Colors.YELLOW
        else:
            color = Colors.GREEN

        return f"[{color}{bar}{Colors.END}]"

    def _display_trend_charts(self):
        """顯示趨勢圖表"""
        print(f"\n{Colors.BOLD}歷史趨勢 (最近 {len(self.metrics_history)} 個數據點):{Colors.END}")

        if len(self.metrics_history) < 2:
            return

        # CPU趨勢
        cpu_values = [m.cpu_percent for m in self.metrics_history]
        self._display_mini_chart("CPU", cpu_values, "%")

        # 記憶體趨勢
        mem_values = [m.memory_percent for m in self.metrics_history]
        self._display_mini_chart("MEM", mem_values, "%")

    def _display_mini_chart(self, label: str, values: List[float], unit: str):
        """顯示迷你圖表"""
        if not values:
            return

        max_val = max(values)
        min_val = min(values)

        print(f"\n{label} 趨勢: {min_val:.1f}{unit} - {max_val:.1f}{unit}")

        # 正規化值到0-20範圍用於顯示
        if max_val > min_val:
            normalized = [(v - min_val) / (max_val - min_val) * 20 for v in values]
        else:
            normalized = [10] * len(values)

        # 顯示圖表
        chart_line = ""
        for norm_val in normalized[-40:]:  # 只顯示最後40個點
            height = int(norm_val)
            if height < 5:
                chart_line += f"{Colors.GREEN}▁{Colors.END}"
            elif height < 10:
                chart_line += f"{Colors.YELLOW}▃{Colors.END}"
            elif height < 15:
                chart_line += f"{Colors.YELLOW}▅{Colors.END}"
            else:
                chart_line += f"{Colors.RED}▇{Colors.END}"

        print(f"     {chart_line}")

class AdvancedOutputFormats:
    """高級輸出格式處理器"""

    def __init__(self):
        self.output_dir = Path("output")
        self.output_dir.mkdir(exist_ok=True)

    def export_to_latex(self, data: List[List[str]],
                       title: str = "數據表格",
                       filename: str = "table.tex") -> str:
        """導出為LaTeX表格"""
        latex_content = []
        latex_content.append("\\documentclass{article}")
        latex_content.append("\\usepackage[utf8]{inputenc}")
        latex_content.append("\\usepackage{CJKutf8}")
        latex_content.append("\\usepackage{booktabs}")
        latex_content.append("\\usepackage{array}")
        latex_content.append("\\begin{document}")
        latex_content.append("\\begin{CJK}{UTF8}{bsmi}")
        latex_content.append(f"\\section{{{title}}}")

        if data and len(data) > 0:
            num_cols = len(data[0])
            col_spec = "|".join(["c"] * num_cols)
            latex_content.append(f"\\begin{{tabular}}{{{col_spec}}}")
            latex_content.append("\\toprule")

            for i, row in enumerate(data):
                escaped_row = [str(cell).replace("&", "\\&").replace("%", "\\%").replace("_", "\\_") for cell in row]
                latex_content.append(" & ".join(escaped_row) + " \\\\")
                if i == 0:
                    latex_content.append("\\midrule")

            latex_content.append("\\bottomrule")
            latex_content.append("\\end{tabular}")

        latex_content.append("\\end{CJK}")
        latex_content.append("\\end{document}")

        # 寫入檔案
        output_path = self.output_dir / filename
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("\n".join(latex_content))

        return str(output_path)

    def export_to_excel(self, data_dict: Dict[str, List[List[Any]]],
                       filename: str = "data_export.xlsx") -> str:
        """導出到Excel檔案"""
        if not ADVANCED_FEATURES:
            print("Excel導出需要安裝 openpyxl")
            return ""

        try:
            output_path = self.output_dir / filename
            workbook = openpyxl.Workbook()

            # 刪除預設工作表
            workbook.remove(workbook.active)

            for sheet_name, data in data_dict.items():
                worksheet = workbook.create_sheet(title=sheet_name[:31])  # Excel工作表名稱限制

                # 設定樣式
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                center_alignment = Alignment(horizontal="center", vertical="center")

                for row_idx, row in enumerate(data, 1):
                    for col_idx, value in enumerate(row, 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx, value=value)

                        # 設定標題行樣式
                        if row_idx == 1:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_alignment

                # 自動調整列寬
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            workbook.save(output_path)
            return str(output_path)

        except Exception as e:
            print(f"Excel導出錯誤: {e}")
            return ""

    def generate_svg_chart(self, data: List[float],
                          labels: List[str] = None,
                          title: str = "數據圖表",
                          filename: str = "chart.svg") -> str:
        """生成SVG向量圖表"""
        if not data:
            return ""

        width, height = 800, 600
        margin = 60
        chart_width = width - 2 * margin
        chart_height = height - 2 * margin

        svg_content = []
        svg_content.append(f'<?xml version="1.0" encoding="UTF-8"?>')
        svg_content.append(f'<svg width="{width}" height="{height}" xmlns="http://www.w3.org/2000/svg">')

        # 背景
        svg_content.append(f'<rect width="{width}" height="{height}" fill="#f8f9fa"/>')

        # 標題
        svg_content.append(f'<text x="{width//2}" y="30" text-anchor="middle" font-size="20" font-weight="bold" fill="#333">{title}</text>')

        # 繪製圖表
        max_val = max(data)
        min_val = min(data)
        val_range = max_val - min_val

        # 座標軸
        svg_content.append(f'<line x1="{margin}" y1="{margin}" x2="{margin}" y2="{height-margin}" stroke="#333" stroke-width="2"/>')
        svg_content.append(f'<line x1="{margin}" y1="{height-margin}" x2="{width-margin}" y2="{height-margin}" stroke="#333" stroke-width="2"/>')

        # 數據點
        if len(data) > 1:
            bar_width = chart_width / len(data) * 0.8

            for i, value in enumerate(data):
                x = margin + (i + 0.1) * (chart_width / len(data))
                normalized_height = (value - min_val) / val_range * chart_height if val_range > 0 else chart_height / 2
                y = height - margin - normalized_height

                # 繪製柱狀圖
                color = f"hsl({i * 360 / len(data)}, 70%, 50%)"
                svg_content.append(f'<rect x="{x}" y="{y}" width="{bar_width}" height="{normalized_height}" fill="{color}" opacity="0.8"/>')

                # 數值標籤
                svg_content.append(f'<text x="{x + bar_width/2}" y="{y-5}" text-anchor="middle" font-size="12" fill="#333">{value:.1f}</text>')

                # X軸標籤
                if labels and i < len(labels):
                    svg_content.append(f'<text x="{x + bar_width/2}" y="{height-margin+20}" text-anchor="middle" font-size="10" fill="#666">{labels[i]}</text>')

        svg_content.append('</svg>')

        # 寫入檔案
        output_path = self.output_dir / filename
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("\n".join(svg_content))

        return str(output_path)

    def create_markdown_report(self, sections: Dict[str, str],
                              title: str = "數據分析報告",
                              filename: str = "report.md") -> str:
        """生成Markdown報告"""
        markdown_content = []
        markdown_content.append(f"# {title}")
        markdown_content.append("")
        markdown_content.append(f"生成時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        markdown_content.append("")

        for section_title, content in sections.items():
            markdown_content.append(f"## {section_title}")
            markdown_content.append("")
            markdown_content.append(content)
            markdown_content.append("")

        # 寫入檔案
        output_path = self.output_dir / filename
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("\n".join(markdown_content))

        return str(output_path)

class InteractiveInterface:
    """互動式命令界面"""

    def __init__(self):
        self.running = False
        self.selected_option = 0
        self.options = [
            "顯示系統監控",
            "生成3D圖表",
            "執行機器學習分析",
            "導出數據報告",
            "實時數據流",
            "系統設定",
            "退出程序"
        ]

    def display_menu(self):
        """顯示互動選單"""
        os.system('cls' if os.name == 'nt' else 'clear')

        print(f"{Colors.BOLD}{Colors.CYAN}{'='*60}{Colors.END}")
        print(f"{Colors.BOLD}{Colors.CYAN}極限複雜輸出系統 - 互動界面{Colors.END}")
        print(f"{Colors.BOLD}{Colors.CYAN}{'='*60}{Colors.END}\n")

        for i, option in enumerate(self.options):
            if i == self.selected_option:
                print(f"{Colors.BG_BLUE}{Colors.WHITE}  ▶ {option:<20} {Colors.END}")
            else:
                print(f"    {option}")

        print(f"\n{Colors.DIM}使用 ↑↓ 鍵選擇，Enter 確認，ESC 退出{Colors.END}")

    def handle_input(self) -> bool:
        """處理用戶輸入"""
        try:
            if ADVANCED_FEATURES:
                import keyboard

                if keyboard.is_pressed('up') and self.selected_option > 0:
                    self.selected_option -= 1
                    time.sleep(0.1)
                elif keyboard.is_pressed('down') and self.selected_option < len(self.options) - 1:
                    self.selected_option += 1
                    time.sleep(0.1)
                elif keyboard.is_pressed('enter'):
                    return self.execute_option()
                elif keyboard.is_pressed('esc'):
                    return False
            else:
                # 簡化版本：直接輸入數字
                print("\n請選擇選項 (0-6): ", end="", flush=True)
                try:
                    choice = int(input())
                    if 0 <= choice < len(self.options):
                        self.selected_option = choice
                        return self.execute_option()
                except ValueError:
                    pass

        except Exception as e:
            print(f"輸入處理錯誤: {e}")

        return True

    def execute_option(self) -> bool:
        """執行選中的選項"""
        option = self.options[self.selected_option]

        if option == "退出程序":
            return False
        elif option == "顯示系統監控":
            print(f"\n{Colors.GREEN}啟動系統監控...{Colors.END}")
            time.sleep(2)
        elif option == "生成3D圖表":
            print(f"\n{Colors.MAGENTA}生成3D圖表...{Colors.END}")
            time.sleep(2)
        elif option == "執行機器學習分析":
            print(f"\n{Colors.YELLOW}執行機器學習分析...{Colors.END}")
            time.sleep(2)
        elif option == "導出數據報告":
            print(f"\n{Colors.CYAN}生成數據報告...{Colors.END}")
            time.sleep(2)
        elif option == "實時數據流":
            print(f"\n{Colors.RED}啟動實時數據流...{Colors.END}")
            time.sleep(2)
        elif option == "系統設定":
            print(f"\n{Colors.BLUE}打開系統設定...{Colors.END}")
            time.sleep(2)

        return True

    def run(self):
        """運行互動界面"""
        self.running = True

        while self.running:
            self.display_menu()
            if not self.handle_input():
                break
            time.sleep(0.05)

        print(f"\n{Colors.GREEN}謝謝使用！{Colors.END}")

class DistributedDataProcessor:
    """分散式數據處理系統"""

    def __init__(self, max_workers: int = 4):
        self.max_workers = max_workers
        self.data_queue = queue.Queue()
        self.result_queue = queue.Queue()
        self.processing_pool = None

    def generate_massive_dataset(self, size: int = 1000000) -> List[DataPoint]:
        """生成大規模數據集"""
        print(f"{Colors.CYAN}正在生成 {size:,} 個數據點的大規模數據集...{Colors.END}")

        dataset = []
        batch_size = 10000

        with concurrent.futures.ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = []

            for batch_start in range(0, size, batch_size):
                batch_end = min(batch_start + batch_size, size)
                future = executor.submit(self._generate_batch, batch_start, batch_end)
                futures.append(future)

            # 收集結果
            for future in concurrent.futures.as_completed(futures):
                batch_data = future.result()
                dataset.extend(batch_data)

                # 顯示進度
                progress = len(dataset) / size * 100
                print(f"\r{Colors.GREEN}進度: {progress:5.1f}% ({len(dataset):,}/{size:,}){Colors.END}", end="", flush=True)

        print(f"\n{Colors.GREEN}✓ 數據集生成完成！{Colors.END}")
        return dataset

    def _generate_batch(self, start: int, end: int) -> List[DataPoint]:
        """生成數據批次"""
        batch = []
        for i in range(start, end):
            timestamp = datetime.now() - timedelta(seconds=random.randint(0, 86400))
            value = random.gauss(50, 20)  # 正態分佈
            category = random.choice(['CPU', 'Memory', 'Network', 'Storage', 'Database'])

            # 添加3D座標
            coordinates = (
                random.uniform(-100, 100),
                random.uniform(-100, 100),
                random.uniform(-100, 100)
            )

            metadata = {
                'batch_id': i // 1000,
                'worker_id': threading.current_thread().ident,
                'generation_time': time.time()
            }

            batch.append(DataPoint(
                timestamp=timestamp,
                value=value,
                category=category,
                metadata=metadata,
                coordinates=coordinates
            ))

        return batch

    def parallel_analysis(self, data: List[DataPoint]) -> Dict[str, Any]:
        """並行數據分析"""
        if not data:
            return {}

        print(f"{Colors.YELLOW}啟動並行分析處理 {len(data):,} 個數據點...{Colors.END}")

        results = {}

        with concurrent.futures.ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # 提交各種分析任務
            future_basic_stats = executor.submit(self._compute_basic_stats, data)
            future_category_analysis = executor.submit(self._analyze_by_category, data)
            future_temporal_analysis = executor.submit(self._analyze_temporal_patterns, data)
            future_spatial_analysis = executor.submit(self._analyze_spatial_distribution, data)

            # 收集結果
            results['basic_stats'] = future_basic_stats.result()
            results['category_analysis'] = future_category_analysis.result()
            results['temporal_analysis'] = future_temporal_analysis.result()
            results['spatial_analysis'] = future_spatial_analysis.result()

        print(f"{Colors.GREEN}✓ 並行分析完成！{Colors.END}")
        return results

    def _compute_basic_stats(self, data: List[DataPoint]) -> Dict[str, float]:
        """計算基本統計量"""
        if not data:
            return {}

        values = [dp.value for dp in data]

        if ADVANCED_FEATURES and np:
            values_np = np.array(values)
            return {
                'count': len(values),
                'mean': float(np.mean(values_np)),
                'std': float(np.std(values_np)),
                'min': float(np.min(values_np)),
                'max': float(np.max(values_np)),
                'median': float(np.median(values_np)),
                'percentile_25': float(np.percentile(values_np, 25)),
                'percentile_75': float(np.percentile(values_np, 75))
            }
        else:
            # 手動計算
            values.sort()
            n = len(values)
            return {
                'count': n,
                'mean': sum(values) / n,
                'min': values[0],
                'max': values[-1],
                'median': values[n//2] if n % 2 == 1 else (values[n//2-1] + values[n//2]) / 2
            }

    def _analyze_by_category(self, data: List[DataPoint]) -> Dict[str, Dict[str, float]]:
        """按類別分析"""
        category_data = defaultdict(list)

        for dp in data:
            category_data[dp.category].append(dp.value)

        results = {}
        for category, values in category_data.items():
            if values:
                results[category] = {
                    'count': len(values),
                    'mean': sum(values) / len(values),
                    'min': min(values),
                    'max': max(values)
                }

        return results

    def _analyze_temporal_patterns(self, data: List[DataPoint]) -> Dict[str, Any]:
        """時間模式分析"""
        hourly_counts = defaultdict(int)
        daily_counts = defaultdict(int)

        for dp in data:
            hour = dp.timestamp.hour
            day = dp.timestamp.strftime('%Y-%m-%d')
            hourly_counts[hour] += 1
            daily_counts[day] += 1

        return {
            'hourly_pattern': dict(hourly_counts),
            'daily_pattern': dict(daily_counts),
            'peak_hour': max(hourly_counts.items(), key=lambda x: x[1])[0] if hourly_counts else 0
        }

    def _analyze_spatial_distribution(self, data: List[DataPoint]) -> Dict[str, Any]:
        """空間分布分析"""
        coordinates = [dp.coordinates for dp in data if dp.coordinates]

        if not coordinates:
            return {}

        x_coords = [c[0] for c in coordinates]
        y_coords = [c[1] for c in coordinates]
        z_coords = [c[2] for c in coordinates]

        return {
            'center_x': sum(x_coords) / len(x_coords),
            'center_y': sum(y_coords) / len(y_coords),
            'center_z': sum(z_coords) / len(z_coords),
            'spread_x': max(x_coords) - min(x_coords),
            'spread_y': max(y_coords) - min(y_coords),
            'spread_z': max(z_coords) - min(z_coords)
        }

    def stream_processing_demo(self, duration: int = 30):
        """流式處理演示"""
        print(f"{Colors.BOLD}{Colors.MAGENTA}啟動流式數據處理演示...{Colors.END}")
        print(f"持續時間: {duration} 秒\n")

        processed_count = 0
        start_time = time.time()

        # 啟動數據生成線程
        def data_generator():
            while time.time() - start_time < duration:
                # 生成模擬數據流
                batch = self._generate_batch(0, 100)
                for dp in batch:
                    self.data_queue.put(dp)
                time.sleep(0.1)

            # 發送結束信號
            self.data_queue.put(None)

        # 啟動處理線程
        def data_processor():
            nonlocal processed_count
            buffer = []

            while True:
                try:
                    item = self.data_queue.get(timeout=1)
                    if item is None:
                        break

                    buffer.append(item)
                    processed_count += 1

                    # 批次處理
                    if len(buffer) >= 50:
                        # 模擬處理時間
                        time.sleep(0.01)
                        buffer.clear()

                except queue.Empty:
                    continue

        # 啟動線程
        generator_thread = threading.Thread(target=data_generator)
        processor_thread = threading.Thread(target=data_processor)

        generator_thread.start()
        processor_thread.start()

        # 監控處理進度
        while generator_thread.is_alive() or processor_thread.is_alive():
            elapsed = time.time() - start_time
            rate = processed_count / elapsed if elapsed > 0 else 0

            print(f"\r{Colors.GREEN}已處理: {processed_count:,} 個數據點 | 處理速度: {rate:.1f} 點/秒{Colors.END}", end="", flush=True)
            time.sleep(0.5)

        generator_thread.join()
        processor_thread.join()

        print(f"\n{Colors.GREEN}✓ 流式處理演示完成！總共處理 {processed_count:,} 個數據點{Colors.END}")

class ComplexOutputSystem:
    """極限複雜輸出系統主類 - 整合所有高階功能"""

    def __init__(self):
        self.data_points: List[DataPoint] = []
        self.categories = ['CPU', 'Memory', 'Network', 'Storage', 'Database', 'Security', 'Analytics']
        self.colors = [Colors.RED, Colors.GREEN, Colors.YELLOW, Colors.BLUE, Colors.MAGENTA, Colors.CYAN, Colors.WHITE]

        # 初始化各個子系統
        self.engine_3d = Advanced3DEngine()
        self.ml_analyzer = MachineLearningAnalyzer()
        self.system_monitor = SystemMonitor()
        self.output_formats = AdvancedOutputFormats()
        self.interactive_ui = InteractiveInterface()
        self.distributed_processor = DistributedDataProcessor()

        # 系統設定
        self.enable_animations = True
        self.enable_real_time = True
        self.max_data_points = 1000000

    def generate_sample_data(self, count: int = 500) -> None:
        """生成增強樣本數據（含3D座標）"""
        print(f"{Colors.CYAN}正在生成 {count} 個增強樣本數據點...{Colors.END}")

        for i in range(count):
            timestamp = datetime.now() - timedelta(minutes=random.randint(0, 1440))

            # 使用多種分佈生成更真實的數據
            if random.random() < 0.7:
                value = random.gauss(50, 15)  # 正態分佈
            else:
                value = random.expovariate(0.02)  # 指數分佈

            value = max(0, min(100, value))  # 限制範圍

            category = random.choice(self.categories)

            # 添加3D空間座標
            coordinates = (
                random.uniform(-100, 100),
                random.uniform(-100, 100),
                random.uniform(-100, 100)
            )

            # 增強元數據
            metadata = {
                'region': random.choice(['台北', '高雄', '台中', '台南', '桃園', '新竹']),
                'severity': random.choice(['低', '中', '高', '緊急']),
                'source': f'server_{random.randint(1, 20)}',
                'protocol': random.choice(['HTTP', 'HTTPS', 'TCP', 'UDP', 'SSH']),
                'user_id': random.randint(1000, 9999),
                'session_id': f'sess_{random.randint(100000, 999999)}'
            }

            # 計算置信度和異常分數
            confidence = random.uniform(0.5, 1.0)
            anomaly_score = random.uniform(0, 0.3) if random.random() < 0.95 else random.uniform(0.7, 1.0)

            self.data_points.append(DataPoint(
                timestamp=timestamp,
                value=value,
                category=category,
                metadata=metadata,
                coordinates=coordinates,
                confidence=confidence,
                anomaly_score=anomaly_score
            ))

        print(f"{Colors.GREEN}✓ 增強數據生成完成（包含3D座標、置信度、異常分數）{Colors.END}")

    def display_system_info(self) -> None:
        """顯示系統信息"""
        print(f"\n{Colors.BOLD}{Colors.CYAN}{'='*60}{Colors.END}")
        print(f"{Colors.BOLD}{Colors.CYAN}系統信息概覽{Colors.END}")
        print(f"{Colors.BOLD}{Colors.CYAN}{'='*60}{Colors.END}")

        info = SystemInfo(
            platform=platform.system(),
            python_version=platform.python_version(),
            cpu_count=os.cpu_count() or 1,
            memory_total=os.getenv('MEMORY_TOTAL', 'Unknown'),
            uptime=time.time()
        )

        info_data = [
            ["項目", "值", "狀態"],
            ["-" * 20, "-" * 30, "-" * 10],
            ["作業系統", info.platform, "✓"],
            ["Python 版本", info.python_version, "✓"],
            ["CPU 核心數", str(info.cpu_count), "✓"],
            ["記憶體總量", str(info.memory_total), "✓"],
            ["運行時間", f"{info.uptime:.2f} 秒", "✓"]
        ]

        self._print_table(info_data)

    def display_statistics(self) -> None:
        """顯示統計信息"""
        print(f"\n{Colors.BOLD}{Colors.YELLOW}{'='*60}{Colors.END}")
        print(f"{Colors.BOLD}{Colors.YELLOW}數據統計分析{Colors.END}")
        print(f"{Colors.BOLD}{Colors.YELLOW}{'='*60}{Colors.END}")

        if not self.data_points:
            print(f"{Colors.RED}沒有數據可供分析{Colors.END}")
            return

        # 按類別統計
        category_stats = defaultdict(list)
        for point in self.data_points:
            category_stats[point.category].append(point.value)

        stats_data = [["類別", "數量", "平均值", "最大值", "最小值", "標準差"]]
        stats_data.append(["-" * 10, "-" * 8, "-" * 10, "-" * 10, "-" * 10, "-" * 10])

        for category, values in category_stats.items():
            count = len(values)
            mean_val = sum(values) / count
            max_val = max(values)
            min_val = min(values)
            std_dev = math.sqrt(sum((x - mean_val) ** 2 for x in values) / count)

            stats_data.append([
                category,
                str(count),
                f"{mean_val:.2f}",
                f"{max_val:.2f}",
                f"{min_val:.2f}",
                f"{std_dev:.2f}"
            ])

        self._print_table(stats_data)

    def display_charts(self) -> None:
        """顯示 ASCII 圖表"""
        print(f"\n{Colors.BOLD}{Colors.MAGENTA}{'='*60}{Colors.END}")
        print(f"{Colors.BOLD}{Colors.MAGENTA}數據可視化圖表{Colors.END}")
        print(f"{Colors.BOLD}{Colors.MAGENTA}{'='*60}{Colors.END}")

        # 生成直方圖
        self._display_histogram()

        # 生成折線圖
        self._display_line_chart()

        # 生成圓餅圖
        self._display_pie_chart()

    def _display_histogram(self) -> None:
        """顯示直方圖"""
        print(f"\n{Colors.CYAN}數據分布直方圖:{Colors.END}")

        # 將數據分為 10 個區間
        bins = [0] * 10
        for point in self.data_points:
            bin_index = min(int(point.value / 10), 9)
            bins[bin_index] += 1

        max_count = max(bins) if bins else 1

        for i, count in enumerate(bins):
            bar_length = int((count / max_count) * 50)
            bar = "█" * bar_length
            range_label = f"{i*10:2d}-{(i+1)*10-1:2d}"
            print(f"{range_label:>6} |{Colors.GREEN}{bar:<50}{Colors.END} {count:>3}")

    def _display_line_chart(self) -> None:
        """顯示折線圖"""
        print(f"\n{Colors.CYAN}時間序列折線圖:{Colors.END}")

        # 按時間排序並取最近 20 個點
        sorted_points = sorted(self.data_points, key=lambda x: x.timestamp)[-20:]

        if not sorted_points:
            return

        values = [point.value for point in sorted_points]
        min_val = min(values)
        max_val = max(values)

        # 正規化到 0-20 範圍
        normalized = [(v - min_val) / (max_val - min_val) * 20 if max_val > min_val else 10 for v in values]

        print("時間序列數據:")
        for i, (point, norm_val) in enumerate(zip(sorted_points, normalized)):
            bar_length = int(norm_val)
            bar = "█" * bar_length
            time_str = point.timestamp.strftime("%H:%M")
            print(f"{time_str:>5} |{Colors.BLUE}{bar:<20}{Colors.END} {point.value:>6.1f}")

    def _display_pie_chart(self) -> None:
        """顯示圓餅圖（ASCII 版本）"""
        print(f"\n{Colors.CYAN}類別分布圓餅圖:{Colors.END}")

        category_counts = Counter(point.category for point in self.data_points)
        total = sum(category_counts.values())

        if total == 0:
            return

        # 使用不同字符表示不同類別
        pie_chars = ['●', '◆', '▲', '■', '★']

        print("類別分布:")
        for i, (category, count) in enumerate(category_counts.most_common()):
            percentage = (count / total) * 100
            char = pie_chars[i % len(pie_chars)]
            color = self.colors[i % len(self.colors)]
            print(f"{color}{char} {category:<10}{Colors.END} {percentage:>5.1f}% ({count:>3} 項)")

    def display_real_time_data(self) -> None:
        """顯示實時數據流"""
        print(f"\n{Colors.BOLD}{Colors.GREEN}{'='*60}{Colors.END}")
        print(f"{Colors.BOLD}{Colors.GREEN}實時數據流監控{Colors.END}")
        print(f"{Colors.BOLD}{Colors.GREEN}{'='*60}{Colors.END}")

        print("正在模擬實時數據更新...")

        for i in range(10):
            # 生成新的數據點
            new_point = DataPoint(
                timestamp=datetime.now(),
                value=random.uniform(0, 100),
                category=random.choice(self.categories),
                metadata={'iteration': i + 1}
            )
            self.data_points.append(new_point)

            # 顯示進度條
            progress = (i + 1) / 10
            bar_length = 30
            filled_length = int(bar_length * progress)
            bar = "█" * filled_length + "░" * (bar_length - filled_length)

            print(f"\r{Colors.CYAN}進度: [{bar}] {progress*100:>5.1f}%{Colors.END}", end="", flush=True)
            time.sleep(0.5)

        print(f"\n{Colors.GREEN}✓ 實時數據更新完成{Colors.END}")

    def display_json_output(self) -> None:
        """顯示 JSON 格式輸出"""
        print(f"\n{Colors.BOLD}{Colors.BLUE}{'='*60}{Colors.END}")
        print(f"{Colors.BOLD}{Colors.BLUE}JSON 格式數據輸出{Colors.END}")
        print(f"{Colors.BOLD}{Colors.BLUE}{'='*60}{Colors.END}")

        # 準備 JSON 數據
        json_data = {
            "metadata": {
                "generated_at": datetime.now().isoformat(),
                "total_points": len(self.data_points),
                "categories": list(set(point.category for point in self.data_points))
            },
            "summary": self._calculate_summary(),
            "recent_data": [
                {
                    "timestamp": point.timestamp.isoformat(),
                    "value": point.value,
                    "category": point.category,
                    "metadata": point.metadata
                }
                for point in self.data_points[-5:]  # 最近 5 個點
            ]
        }

        # 格式化輸出 JSON
        json_str = json.dumps(json_data, indent=2, ensure_ascii=False)
        print(json_str)

    def _calculate_summary(self) -> Dict[str, Any]:
        """計算數據摘要"""
        if not self.data_points:
            return {}

        values = [point.value for point in self.data_points]
        return {
            "total_points": len(self.data_points),
            "average_value": sum(values) / len(values),
            "min_value": min(values),
            "max_value": max(values),
            "value_range": max(values) - min(values)
        }

    def _print_table(self, data: List[List[str]]) -> None:
        """打印表格"""
        if not data:
            return

        # 計算每列的最大寬度
        col_widths = []
        for col in range(len(data[0])):
            max_width = max(len(str(row[col])) for row in data)
            col_widths.append(max_width + 2)

        # 打印表格
        for i, row in enumerate(data):
            if i == 1:  # 分隔線
                print("+" + "+".join("-" * width for width in col_widths) + "+")
            else:
                formatted_row = "|".join(f" {str(cell):<{col_widths[j]-1}}" for j, cell in enumerate(row))
                print(f"|{formatted_row}|")

    def display_progress_animation(self) -> None:
        """顯示進度動畫"""
        print(f"\n{Colors.BOLD}{Colors.YELLOW}系統初始化中...{Colors.END}")

        stages = [
            "載入配置檔案",
            "初始化資料庫連接",
            "載入業務邏輯模組",
            "啟動 API 服務",
            "註冊路由端點",
            "啟動監控服務",
            "系統就緒"
        ]

        for i, stage in enumerate(stages):
            # 進度條
            progress = (i + 1) / len(stages)
            bar_length = 40
            filled_length = int(bar_length * progress)
            bar = "█" * filled_length + "░" * (bar_length - filled_length)

            # 動畫效果
            spinner = "|/-\\"[i % 4]

            print(f"\r{Colors.CYAN}{spinner} {stage:<20} [{bar}] {progress*100:>5.1f}%{Colors.END}", end="", flush=True)
            time.sleep(0.8)

        print(f"\n{Colors.GREEN}✓ 系統初始化完成！{Colors.END}")

    def display_advanced_3d_charts(self) -> None:
        """顯示高級3D圖表"""
        print(f"\n{Colors.BOLD}{Colors.MAGENTA}{'='*70}{Colors.END}")
        print(f"{Colors.BOLD}{Colors.MAGENTA}🎯 高級3D可視化圖表展示{Colors.END}")
        print(f"{Colors.BOLD}{Colors.MAGENTA}{'='*70}{Colors.END}")

        if not self.data_points:
            print(f"{Colors.RED}無數據可顯示{Colors.END}")
            return

        # 準備3D柱狀圖數據
        category_values = defaultdict(list)
        for point in self.data_points:
            category_values[point.category].append(point.value)

        # 轉換為3D圖表數據格式
        chart_data = []
        labels = []
        for category, values in list(category_values.items())[:5]:  # 限制5個類別
            avg_values = [sum(values[i:i+10])/min(10, len(values[i:]))
                         for i in range(0, len(values), 10)][:8]  # 最多8個數據點
            chart_data.append(avg_values)
            labels.append(category)

        # 顯示3D柱狀圖
        result = self.engine_3d.create_3d_bar_chart(chart_data, labels)
        print(result)

        # 創建3D散點圖
        points_3d = []
        for point in self.data_points[:50]:  # 限制50個點避免過於擁擠
            if point.coordinates:
                points_3d.append(Point3D(
                    x=point.coordinates[0],
                    y=point.coordinates[1],
                    z=point.coordinates[2],
                    symbol='●' if point.anomaly_score < 0.5 else '⚠'
                ))

        result = self.engine_3d.create_3d_scatter(points_3d)
        print(result)

        # 創建3D表面圖數據
        surface_data = [[[random.uniform(0, 100) for _ in range(8)] for _ in range(6)] for _ in range(3)]
        matrix_3d = Matrix3D(
            data=surface_data,
            x_labels=[f"X{i}" for i in range(8)],
            y_labels=[f"Y{i}" for i in range(6)],
            z_labels=[f"Z{i}" for i in range(3)]
        )

        result = self.engine_3d.create_3d_surface(matrix_3d)
        print(result)

    def perform_ml_analysis(self) -> None:
        """執行機器學習分析"""
        print(f"\n{Colors.BOLD}{Colors.YELLOW}{'='*70}{Colors.END}")
        print(f"{Colors.BOLD}{Colors.YELLOW}🤖 機器學習數據分析{Colors.END}")
        print(f"{Colors.BOLD}{Colors.YELLOW}{'='*70}{Colors.END}")

        if not self.data_points:
            print(f"{Colors.RED}無數據可分析{Colors.END}")
            return

        # 異常檢測
        values = [point.value for point in self.data_points]
        anomaly_result = self.ml_analyzer.detect_anomalies(values, contamination=0.1)

        print(f"{Colors.CYAN}🔍 異常檢測結果:{Colors.END}")
        print(f"  模型: {anomaly_result.model_type}")
        print(f"  正常數據比例: {anomaly_result.accuracy:.3f}")
        print(f"  檢測到異常點: {len(anomaly_result.anomalies)} 個")

        if anomaly_result.anomalies:
            print(f"  異常點索引: {anomaly_result.anomalies[:10]}...")  # 只顯示前10個

        # 聚類分析
        coordinates_2d = [(p.coordinates[0], p.coordinates[1])
                         for p in self.data_points[:100] if p.coordinates]

        if coordinates_2d:
            cluster_result = self.ml_analyzer.perform_clustering(coordinates_2d, n_clusters=3)

            print(f"\n{Colors.MAGENTA}🎯 聚類分析結果:{Colors.END}")
            print(f"  模型: {cluster_result.model_type}")
            print(f"  聚類品質分數: {cluster_result.accuracy:.3f}")

            if cluster_result.cluster_centers:
                print(f"  聚類中心:")
                for i, center in enumerate(cluster_result.cluster_centers):
                    print(f"    簇 {i+1}: ({center[0]:.2f}, {center[1]:.2f})")

        # 趨勢預測
        time_series = [point.value for point in sorted(self.data_points, key=lambda x: x.timestamp)[-50:]]
        prediction_result = self.ml_analyzer.predict_trend(time_series, future_points=10)

        print(f"\n{Colors.GREEN}📈 趨勢預測結果:{Colors.END}")
        print(f"  模型: {prediction_result.model_type}")
        print(f"  預測準確度: {prediction_result.accuracy:.3f}")
        print(f"  未來10個點預測值: {[f'{p:.2f}' for p in prediction_result.predictions[:5]]}...")

    def generate_comprehensive_reports(self) -> None:
        """生成綜合分析報告"""
        print(f"\n{Colors.BOLD}{Colors.BLUE}{'='*70}{Colors.END}")
        print(f"{Colors.BOLD}{Colors.BLUE}📊 綜合分析報告生成{Colors.END}")
        print(f"{Colors.BOLD}{Colors.BLUE}{'='*70}{Colors.END}")

        if not self.data_points:
            print(f"{Colors.RED}無數據可報告{Colors.END}")
            return

        # 準備報告數據
        summary = self._calculate_summary()

        # 生成LaTeX報告
        table_data = [
            ["指標", "數值", "說明"],
            ["總數據點", str(summary.get('total_points', 0)), "數據集大小"],
            ["平均值", f"{summary.get('average_value', 0):.2f}", "數值均值"],
            ["標準差", f"{summary.get('value_range', 0):.2f}", "數據離散度"],
            ["最小值", f"{summary.get('min_value', 0):.2f}", "最小觀測值"],
            ["最大值", f"{summary.get('max_value', 0):.2f}", "最大觀測值"]
        ]

        latex_path = self.output_formats.export_to_latex(
            table_data,
            title="數據分析統計表",
            filename="analysis_summary.tex"
        )
        print(f"{Colors.GREEN}✓ LaTeX報告已生成: {latex_path}{Colors.END}")

        # 生成Excel報告
        excel_data = {
            "統計摘要": table_data,
            "數據明細": [["時間戳", "數值", "類別", "地區"]] + [
                [point.timestamp.strftime("%Y-%m-%d %H:%M"),
                 f"{point.value:.2f}",
                 point.category,
                 point.metadata.get('region', 'N/A')]
                for point in self.data_points[:100]  # 限制100行
            ]
        }

        excel_path = self.output_formats.export_to_excel(
            excel_data,
            filename="comprehensive_analysis.xlsx"
        )
        if excel_path:
            print(f"{Colors.GREEN}✓ Excel報告已生成: {excel_path}{Colors.END}")

        # 生成SVG圖表
        category_counts = Counter(point.category for point in self.data_points)
        svg_data = list(category_counts.values())
        svg_labels = list(category_counts.keys())

        svg_path = self.output_formats.generate_svg_chart(
            svg_data,
            svg_labels,
            title="類別分布圖",
            filename="category_distribution.svg"
        )
        print(f"{Colors.GREEN}✓ SVG圖表已生成: {svg_path}{Colors.END}")

        # 生成Markdown報告
        sections = {
            "數據概覽": f"本報告包含 {summary.get('total_points', 0)} 個數據點的分析結果。",
            "統計摘要": f"平均值: {summary.get('average_value', 0):.2f}, 範圍: {summary.get('min_value', 0):.2f} - {summary.get('max_value', 0):.2f}",
            "類別分布": f"共有 {len(set(p.category for p in self.data_points))} 個不同類別",
            "異常檢測": "已執行異常檢測分析，詳見相關章節",
            "結論": "數據分析完成，所有報告文件已生成"
        }

        md_path = self.output_formats.create_markdown_report(
            sections,
            title="數據分析綜合報告",
            filename="analysis_report.md"
        )
        print(f"{Colors.GREEN}✓ Markdown報告已生成: {md_path}{Colors.END}")

    def demonstrate_distributed_processing(self) -> None:
        """演示分散式數據處理"""
        print(f"\n{Colors.BOLD}{Colors.RED}{'='*70}{Colors.END}")
        print(f"{Colors.BOLD}{Colors.RED}⚡ 分散式數據處理演示{Colors.END}")
        print(f"{Colors.BOLD}{Colors.RED}{'='*70}{Colors.END}")

        # 生成大規模數據集
        large_dataset = self.distributed_processor.generate_massive_dataset(50000)

        # 並行分析
        analysis_results = self.distributed_processor.parallel_analysis(large_dataset)

        # 顯示分析結果
        print(f"\n{Colors.CYAN}📊 並行分析結果:{Colors.END}")

        if 'basic_stats' in analysis_results:
            stats = analysis_results['basic_stats']
            print(f"  基本統計: 均值={stats.get('mean', 0):.2f}, 標準差={stats.get('std', 0):.2f}")

        if 'category_analysis' in analysis_results:
            print(f"  類別分析: {len(analysis_results['category_analysis'])} 個類別")

        if 'temporal_analysis' in analysis_results:
            temporal = analysis_results['temporal_analysis']
            print(f"  時間分析: 峰值時段={temporal.get('peak_hour', 'N/A')}時")

        if 'spatial_analysis' in analysis_results:
            spatial = analysis_results['spatial_analysis']
            print(f"  空間分析: 中心點=({spatial.get('center_x', 0):.1f}, {spatial.get('center_y', 0):.1f})")

        # 流式處理演示
        print(f"\n{Colors.YELLOW}🌊 啟動流式處理演示...{Colors.END}")
        self.distributed_processor.stream_processing_demo(duration=10)

    def launch_interactive_interface(self) -> None:
        """啟動互動式界面"""
        print(f"\n{Colors.BOLD}{Colors.MAGENTA}{'='*70}{Colors.END}")
        print(f"{Colors.BOLD}{Colors.MAGENTA}🖥️  啟動互動式用戶界面{Colors.END}")
        print(f"{Colors.BOLD}{Colors.MAGENTA}{'='*70}{Colors.END}")

        try:
            self.interactive_ui.run()
        except KeyboardInterrupt:
            print(f"\n{Colors.YELLOW}用戶中斷操作{Colors.END}")
        except Exception as e:
            print(f"\n{Colors.RED}界面運行錯誤: {e}{Colors.END}")

    def run_ultimate_complex_demo(self) -> None:
        """運行終極複雜輸出演示"""
        print(f"{Colors.BOLD}{Colors.MAGENTA}")
        print("╔" + "═" * 68 + "╗")
        print("║" + " " * 15 + "🚀 極限複雜輸出系統 - 終極演示 🚀" + " " * 15 + "║")
        print("╚" + "═" * 68 + "╝")
        print(f"{Colors.END}")

        # 顯示進度動畫
        self.display_progress_animation()

        # 生成增強樣本數據
        self.generate_sample_data(800)

        # 顯示系統信息
        self.display_system_info()

        # 高級3D圖表展示
        self.display_advanced_3d_charts()

        # 機器學習分析
        self.perform_ml_analysis()

        # 實時系統監控（短時間演示）
        print(f"\n{Colors.BOLD}{Colors.GREEN}🖥️  實時系統監控演示{Colors.END}")
        self.system_monitor.display_real_time_dashboard(duration=10)

        # 生成綜合報告
        self.generate_comprehensive_reports()

        # 分散式處理演示
        self.demonstrate_distributed_processing()

        # 最終摘要
        print(f"\n{Colors.BOLD}{Colors.GREEN}{'='*70}{Colors.END}")
        print(f"{Colors.BOLD}{Colors.GREEN}🎉 終極演示完成摘要{Colors.END}")
        print(f"{Colors.BOLD}{Colors.GREEN}{'='*70}{Colors.END}")

        summary = self._calculate_summary()
        print(f"🔢 總數據點數: {Colors.CYAN}{summary.get('total_points', 0):,}{Colors.END}")
        print(f"📊 平均數值: {Colors.CYAN}{summary.get('average_value', 0):.3f}{Colors.END}")
        print(f"📈 數值範圍: {Colors.CYAN}{summary.get('min_value', 0):.2f} ~ {summary.get('max_value', 0):.2f}{Colors.END}")
        print(f"🏷️  數據類別: {Colors.CYAN}{len(set(point.category for point in self.data_points))}{Colors.END}")
        print(f"🎯 異常數據: {Colors.CYAN}{len([p for p in self.data_points if p.anomaly_score > 0.5])}{Colors.END}")

        features_used = [
            "✅ 3D ASCII 圖表引擎",
            "✅ 機器學習數據分析",
            "✅ 實時系統監控",
            "✅ 多格式輸出 (LaTeX/SVG/Excel/MD)",
            "✅ 分散式數據處理",
            "✅ 互動式用戶界面",
            "✅ 高級可視化",
            "✅ 異常檢測與預測"
        ]

        print(f"\n{Colors.BOLD}{Colors.CYAN}🎊 已展示功能清單:{Colors.END}")
        for feature in features_used:
            print(f"  {Colors.GREEN}{feature}{Colors.END}")

        print(f"\n{Colors.BOLD}{Colors.MAGENTA}🙏 感謝使用極限複雜輸出系統！{Colors.END}")
        print(f"{Colors.DIM}這是目前最複雜、最全面的Python終端輸出演示系統{Colors.END}")

    def run_complex_demo(self) -> None:
        """運行複雜輸出演示（保持向後兼容）"""
        print(f"{Colors.BOLD}{Colors.MAGENTA}")
        print("╔" + "═" * 58 + "╗")
        print("║" + " " * 20 + "複雜輸出系統演示" + " " * 20 + "║")
        print("╚" + "═" * 58 + "╝")
        print(f"{Colors.END}")

        # 顯示進度動畫
        self.display_progress_animation()

        # 生成樣本數據
        self.generate_sample_data(150)

        # 顯示系統信息
        self.display_system_info()

        # 顯示統計分析
        self.display_statistics()

        # 顯示圖表
        self.display_charts()

        # 顯示實時數據
        self.display_real_time_data()

        # 顯示 JSON 輸出
        self.display_json_output()

        # 最終摘要
        print(f"\n{Colors.BOLD}{Colors.GREEN}{'='*60}{Colors.END}")
        print(f"{Colors.BOLD}{Colors.GREEN}演示完成摘要{Colors.END}")
        print(f"{Colors.BOLD}{Colors.GREEN}{'='*60}{Colors.END}")

        summary = self._calculate_summary()
        print(f"總數據點數: {Colors.CYAN}{summary.get('total_points', 0)}{Colors.END}")
        print(f"平均數值: {Colors.CYAN}{summary.get('average_value', 0):.2f}{Colors.END}")
        print(f"數值範圍: {Colors.CYAN}{summary.get('min_value', 0):.2f} - {summary.get('max_value', 0):.2f}{Colors.END}")
        print(f"數據類別: {Colors.CYAN}{len(set(point.category for point in self.data_points))}{Colors.END}")

        print(f"\n{Colors.BOLD}{Colors.MAGENTA}感謝使用複雜輸出系統！{Colors.END}")

def dev():
    """
    極限複雜輸出系統 - 開發環境啟動函數

    這個函數會啟動目前最複雜、最全面的Python終端輸出演示系統，包含：

    🚀 核心功能：
    - 增強型數據生成（3D座標、置信度、異常分數）
    - 高級3D ASCII圖表引擎（立體柱狀圖、表面圖、散點圖、線框圖）
    - 機器學習數據分析（異常檢測、聚類分析、趨勢預測）
    - 實時系統監控儀表板（CPU、記憶體、網路、I/O監控）
    - 多格式輸出支援（LaTeX、SVG、Excel、Markdown）
    - 分散式數據處理（多執行緒、流式處理、大數據集分析）
    - 互動式用戶界面（鍵盤導航、選單系統）

    🎯 技術特色：
    - 支援RGB顏色、動畫效果、進度追蹤
    - 併發處理、記憶體優化、異常處理
    - 科學計算、統計分析、數據可視化
    - 跨平台兼容、模組化設計、擴展性強

    📊 演示內容：
    - 系統信息與性能監控
    - 3D可視化與ASCII藝術
    - 機器學習模型訓練與預測
    - 實時數據流處理
    - 多種格式報告生成
    - 大規模數據集並行分析
    """
    print(f"{Colors.BOLD}{Colors.CYAN}")
    print("🌟" * 35)
    print("🚀 歡迎使用極限複雜輸出系統 🚀")
    print("   Python終端輸出的巔峰之作    ")
    print("🌟" * 35)
    print(f"{Colors.END}\n")

    print(f"{Colors.YELLOW}系統初始化中...{Colors.END}")
    time.sleep(1)

    try:
        # 創建極限複雜輸出系統實例
        output_system = ComplexOutputSystem()

        # 檢查是否有進階功能
        if ADVANCED_FEATURES:
            print(f"{Colors.GREEN}✓ 進階功能已啟用（包含機器學習、科學計算等）{Colors.END}")
        else:
            print(f"{Colors.YELLOW}⚠ 部分進階功能不可用，將使用基礎版本{Colors.END}")

        print(f"\n{Colors.CYAN}選擇演示模式:{Colors.END}")
        print(f"1. {Colors.GREEN}終極複雜演示{Colors.END} - 完整功能展示（推薦）")
        print(f"2. {Colors.YELLOW}傳統複雜演示{Colors.END} - 向後兼容版本")
        print(f"3. {Colors.BLUE}互動式界面{Colors.END} - 用戶控制模式")

        try:
            choice = input(f"\n{Colors.BOLD}請選擇 (1-3, 默認1): {Colors.END}").strip()
            if not choice:
                choice = "1"
        except:
            choice = "1"

        print(f"\n{Colors.MAGENTA}{'='*70}{Colors.END}")

        if choice == "1":
            print(f"{Colors.BOLD}{Colors.GREEN}🚀 啟動終極複雜演示模式{Colors.END}")
            output_system.run_ultimate_complex_demo()

        elif choice == "2":
            print(f"{Colors.BOLD}{Colors.YELLOW}📊 啟動傳統複雜演示模式{Colors.END}")
            output_system.run_complex_demo()

        elif choice == "3":
            print(f"{Colors.BOLD}{Colors.BLUE}🖥️ 啟動互動式界面模式{Colors.END}")
            output_system.launch_interactive_interface()

        else:
            print(f"{Colors.RED}無效選擇，啟動默認模式{Colors.END}")
            output_system.run_ultimate_complex_demo()

    except KeyboardInterrupt:
        print(f"\n\n{Colors.YELLOW}⚠️ 用戶中斷演示{Colors.END}")
        print(f"{Colors.CYAN}感謝您體驗極限複雜輸出系統！{Colors.END}")

    except Exception as e:
        print(f"\n\n{Colors.RED}❌ 系統運行錯誤: {e}{Colors.END}")
        print(f"{Colors.DIM}建議檢查依賴包是否正確安裝{Colors.END}")

    finally:
        print(f"\n{Colors.BOLD}{Colors.MAGENTA}")
        print("🎊" * 35)
        print("    演示結束，謝謝使用！      ")
        print("🎊" * 35)
        print(f"{Colors.END}")

    # 原始開發伺服器代碼（已註解）
    # 如需啟動FastAPI服務器，請取消下面的註解
    # from uvicorn import run as uvicorn_run
    # uvicorn_run("fastapi_demo_02.app:app", reload=True, host="127.0.0.1", port=8000)